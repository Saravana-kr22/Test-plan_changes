import re
import openpyxl
from openpyxl import load_workbook
import sqlite3

excel_file = f"../test_plan_change.xlsx"

def write_excel(test_plan):


    try:
        workbook = load_workbook(excel_file)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet_names = workbook.sheetnames


    if "All_TC_Details" in sheet_names:
            workbook.remove(workbook["All_TC_Details"])
            workbook.create_sheet("All_TC_Details", 0)
            sheet1 = workbook["All_TC_Details"]

    else:
            sheet1 = workbook.active
            sheet1.title = "All_TC_Details"

    head = ["S.no" ,"Cluster Name","Test Case Name","Test Case ID"," Test Plan "]
    sheet1.append(head)
    conn = sqlite3.connect('all_tc_details.db')
    data = conn.execute('''SELECT * FROM Alltcdetails''')
    i =1
    for line in data:
        line = (i,)+line
        sheet1.append(line)
        i+=1
    column_widths = {'A': 10, 'B': 20, 'C': 50 ,'D':30,'E':30}  # Specify the column widths as desired

    for column, width in column_widths.items():
                sheet1.column_dimensions[column].width = width
    workbook.save(excel_file)
    tc_changes = []

    try:
        data = conn.execute('''SELECT * FROM Test_plan_changes''')
        for line in data:
            tc_changes.append(line)
    except:
         pass
    

    summary_changes = []
    try:
        data = conn.execute('''SELECT * FROM Summary_changes''')
        for line in data:
            summary_changes.append(line)
    except:
         pass
    
    conn.close()
    if "Test_Summary_Changes" not in sheet_names:

        sheet = workbook.create_sheet("Test_Summary_Changes",1)
        sheet.append(["Date of Run"	, "Cluster Name",	"Test Case Name",	"Test Case ID"])

    else:
        sheet = workbook["Test_Summary_Changes"]

    for i in range(len(summary_changes)):
        sheet.insert_rows(2)
        print(1,summary_changes[i])
        for j, value in enumerate(summary_changes[i]):                                                                                 
            sheet.cell(row=2, column=j + 1, value=str(value))
    column_widths = {'A': 10, 'B': 20, 'C': 50 ,'D':30,'E':30}  # Specify the column widths as desired

    for column, width in column_widths.items():
                sheet1.column_dimensions[column].width = width
    
    workbook.save(excel_file)

    if "Test_plan_Changes" not in sheet_names:

        sheet = workbook.create_sheet("Test_plan_Changes",2)
        sheet.append(["Date of Run"," Commit","Cluster/Testcase","Changes"])
    else:
        sheet = workbook["Test_plan_Changes"]

    for i in range(len(tc_changes)):
        sheet.insert_rows(2) 
        print(2,tc_changes[i])
        for j, value in enumerate(tc_changes[i]):                                                                                
            sheet.cell(row=2, column=j + 1, value=str(value))

    column_widths = {'A': 10, 'B': 20, 'C': 50 ,'D':30,'E':30}  # Specify the column widths as desired

    for column, width in column_widths.items():
                sheet1.column_dimensions[column].width = width

    workbook.save(excel_file)
         
    cluster_codes = []

    for test in test_plan:
            tcid = test_plan[test][0]["Test Case ID"]
            sh = re.search(r'-(.*?)-', tcid)
            code = sh.group(1)
            if code == 'LOWPOWER':
                code = 'MC'

            cluster_codes.append(code)

    for code in cluster_codes:
            if code in sheet_names:
                workbook.remove(workbook[code])
                workbook.create_sheet(code, sheet_names.index(code))
                

            else:
                workbook.create_sheet(code)

    tcsummary(workbook, test_plan)
    workbook.save(excel_file)



def tcsummary(workbook, updated_data):
    clusters = list(updated_data.keys())
    for cluster in clusters:
        tcids = updated_data[cluster][0]["Test Case ID"]
        sh = re.search(r'-(.*?)-', tcids)
        code = sh.group(1)
        if code == 'LOWPOWER':
            code = 'MC'
        sheet = workbook[code]
        sheet.append([cluster])
        sheet.append([""])

        testcases = updated_data[cluster]
        for testcase in testcases:
            testcase_name = testcase["Test Case Name"]
            print(testcase_name)
            sheet.append([testcase_name])
            sheet.append([""])
            sheet.append(["Purpose"])
            sheet.append([testcase["Purpose"]])
            sheet.append([""])
            sheet.append(["PICS"])
            for pics in testcase["PICS"]:
                sheet.append([pics])
            sheet.append([""])
            sheet.append(["Pre-condition"])
            if testcase["Pre-condition"] =="Nil":
                sheet.append(["Nil"])
            else:
                head = list(testcase["Pre-condition"].keys())
                sheet.append(head)
                for i in range(len(list(testcase["Pre-condition"].values())[0])):
                    val =[]
                    for key, value in testcase["Pre-condition"].items():
                        if i < len(value):
                            val.append(value[i]) 
                    sheet.append(val)
            sheet.append([""])
            sheet.append(["Test Procedure"])
            keys = list(testcase["Test Procedure"].keys())
            sheet.append(keys)
            for i in range(len(list(testcase["Test Procedure"].values())[0])):
                val =[]
                for key, value in testcase["Test Procedure"].items():
                    if i < len(value):
                        val.append(value[i]) 
                sheet.append(val)
            sheet.append([""])
            sheet.append([""])
            sheet.append([""])
            
            workbook.save(excel_file)
